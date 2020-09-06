import openpyxl


def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return(sheet.max_row)


def getColCount(file, sheetname):
    workboook = openpyxl.load_workbook(file)
    sheet = workboook[sheetname]
    return(sheet.max_column)


def readData(file, sheetname, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(row=rownum, column=columnno).value



def writeData(file, sheetname, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(row=rownum, column=columnno).value
    workbook.save(file)

